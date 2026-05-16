"""Manage data sources: view what's loaded, add new CSV/JSON sources, load Annex I Excel."""
from __future__ import annotations

# --- Ensure project root is on sys.path (Streamlit/Railway compat) -----
import sys
from pathlib import Path
_ROOT = Path(__file__).resolve().parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))
# ----------------------------------------------------------------------

import io
import json
import time

import pandas as pd
import streamlit as st

from db.connection import execute, run_query, bulk_insert_values

st.set_page_config(page_title="Data sources", page_icon="⚙️", layout="wide")
st.title("⚙️ Data sources")
st.caption(
    "Every dataset used by the sandbox is tracked here for reproducibility. "
    "Each source has a name, version, and load timestamp."
)


# ----------------------------------------------------------------------
# Current sources table
# ----------------------------------------------------------------------
st.subheader("Currently registered sources")
try:
    rows = run_query(
        """
        SELECT  id, source_type, source_name, version, loaded_at,
                row_count, is_active, notes
        FROM    data_sources
        ORDER BY loaded_at DESC
        """
    )
    if not rows:
        st.info("No sources loaded yet. Use the upload sections below.")
    else:
        df = pd.DataFrame(rows)
        st.dataframe(df, hide_index=True, use_container_width=True)
except Exception as exc:
    st.error(f"Could not read data_sources: {exc}")
    st.info("Verify DATABASE_URL is set and `python -m db.init_db` has been run.")


st.divider()

# ----------------------------------------------------------------------
# Section 1: Upload the Annex I Excel
# ----------------------------------------------------------------------
st.subheader("📚 Load Annex I from DG TRADE Excel")
st.caption(
    "Excel with columns `ID`, `PARENT_ID`, `CODE`, `LABEL` on the "
    "**Export Worksheet** sheet — as published by DG TRADE."
)

col_file, col_version = st.columns([2, 1])
with col_file:
    excel_file = st.file_uploader(
        "Annex I Excel (.xlsx)",
        type=["xlsx"],
        key="annex_upload",
    )
with col_version:
    version_label = st.text_input(
        "Version label",
        value="2024-09",
        placeholder="2024-09",
        key="annex_version",
        help="The DG TRADE Excel currently available is the September 2024 version. "
             "Change this when you load a newer version.",
    )

# Button is ALWAYS rendered, just disabled until prerequisites are met.
ready = (excel_file is not None) and bool(version_label.strip())
load_clicked = st.button(
    "🚀 Load Annex I into database" if ready else "Upload a file and enter a version first",
    type="primary",
    disabled=not ready,
    use_container_width=True,
    key="load_annex_btn",
)

if load_clicked:
    progress = st.progress(0, text="Reading Excel...")
    status = st.empty()
    try:
        import openpyxl

        # Use getvalue() — it does NOT consume the stream like read() does.
        file_bytes = excel_file.getvalue()
        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes),
            data_only=True,
            read_only=True,
        )
        if "Export Worksheet" not in wb.sheetnames:
            st.error(
                f"Sheet 'Export Worksheet' not found. "
                f"Available sheets: {wb.sheetnames}"
            )
            st.stop()
        ws = wb["Export Worksheet"]
        raw_rows = list(ws.iter_rows(min_row=2, values_only=True))
        progress.progress(20, text=f"Parsed {len(raw_rows)} rows from Excel.")
        status.info(f"📖 Read {len(raw_rows):,} rows from {excel_file.name}.")

        # ---- Register source row ------------------------------------
        progress.progress(30, text="Registering source...")
        version_str = version_label.strip()
        existing = run_query(
            "SELECT id FROM data_sources WHERE source_type='annex_i' AND version=:v",
            {"v": version_str},
        )
        if existing:
            source_id = existing[0]["id"]
            execute(
                "DELETE FROM annex_i_items WHERE source_id=:sid",
                {"sid": source_id},
            )
            status.info(f"♻️ Replacing existing source #{source_id} ({version_str}).")
        else:
            inserted = run_query(
                """
                INSERT INTO data_sources (source_type, source_name, version, row_count, notes)
                VALUES ('annex_i', :n, :v, :rc, :notes)
                RETURNING id
                """,
                {
                    "n": f"EU Annex I Dual-Use ({version_str})",
                    "v": version_str,
                    "rc": len(raw_rows),
                    "notes": f"Uploaded via UI: {excel_file.name}",
                },
            )
            source_id = inserted[0]["id"]
            execute(
                """
                UPDATE data_sources SET is_active = FALSE
                WHERE  source_type = 'annex_i' AND id <> :sid
                """,
                {"sid": source_id},
            )
            status.info(f"✨ Registered new source #{source_id} ({version_str}).")

        # ---- Build batch --------------------------------------------
        progress.progress(50, text="Preparing rows for insert...")
        batch = []
        for r in raw_rows:
            item_id, parent_id, code, label = r
            if item_id is None or code is None:
                continue
            code_str = str(code).strip()
            category = code_str[0] if code_str and code_str[0].isdigit() else None
            subgroup = code_str[1] if (category and len(code_str) >= 2 and code_str[1] in "ABCDE") else None
            depth = code_str.count(".")
            if parent_id in ("", None):
                pid_val = None
            else:
                try:
                    pid_val = int(parent_id)
                except (TypeError, ValueError):
                    pid_val = None
            batch.append({
                "id": int(item_id),
                "parent_id": pid_val,
                "code": code_str,
                "label": str(label) if label is not None else "",
                "category": category,
                "subgroup": subgroup,
                "depth": depth,
                "source_id": source_id,
            })

        # ---- Fast bulk insert via psycopg2.execute_values -----------
        progress.progress(70, text=f"Inserting {len(batch):,} rows into Postgres...")
        t0 = time.time()
        bulk_insert_values(
            table="annex_i_items",
            columns=["id", "parent_id", "code", "label",
                     "category", "subgroup", "depth", "source_id"],
            rows=batch,
            page_size=500,
        )
        elapsed = time.time() - t0

        execute(
            "UPDATE data_sources SET row_count = :rc WHERE id = :sid",
            {"rc": len(batch), "sid": source_id},
        )

        progress.progress(100, text="Done.")
        status.success(
            f"✓ Loaded {len(batch):,} Annex I rows in {elapsed:.1f}s "
            f"(source #{source_id}, version {version_str})."
        )
        time.sleep(1.5)
        st.rerun()

    except Exception as exc:
        progress.empty()
        status.empty()
        st.error(f"Load failed: {exc}")
        st.exception(exc)


st.divider()

# ----------------------------------------------------------------------
# Section 2: Upload Dutch (or other language) regulation text
# ----------------------------------------------------------------------
st.subheader("📄 Load regulation text (Dual_use.txt — Dutch / EN / FR / DE)")
st.caption(
    "Plaintext export of the consolidated EU regulation from EUR-Lex. "
    "Extracts ECN code → first-line label pairs and stores them as "
    "`manual_entries` with source_type `annex_i_text`. "
    "Searchable from the **Search product** page alongside the structured Excel."
)

col_txt, col_lang, col_ver = st.columns([2, 1, 1])
with col_txt:
    txt_file = st.file_uploader(
        "Regulation text file (.txt)",
        type=["txt"],
        key="regulation_txt_upload",
    )
with col_lang:
    txt_language = st.selectbox(
        "Language",
        options=["NL", "EN", "FR", "DE", "other"],
        index=0,
        key="regulation_txt_lang",
    )
with col_ver:
    txt_version = st.text_input(
        "Version",
        value="2025-11",
        placeholder="2025-11",
        key="regulation_txt_version",
    )

txt_ready = (txt_file is not None) and bool(txt_version.strip())
txt_clicked = st.button(
    "🚀 Parse and load regulation text" if txt_ready
    else "Upload a .txt and enter a version first",
    type="primary",
    disabled=not txt_ready,
    use_container_width=True,
    key="load_txt_btn",
)

if txt_clicked:
    progress = st.progress(0, text="Reading file...")
    status = st.empty()
    try:
        from services.regulation_parser import parse_eu_regulation_txt

        # Decode bytes → string (try UTF-8 first, fall back to latin-1)
        raw = txt_file.getvalue()
        try:
            text_content = raw.decode("utf-8")
        except UnicodeDecodeError:
            text_content = raw.decode("latin-1")
        progress.progress(20, text=f"Read {len(text_content):,} chars.")

        # Parse
        progress.progress(40, text="Extracting ECN entries...")
        entries = parse_eu_regulation_txt(text_content)
        if not entries:
            st.warning("No ECN-style entries found in the file. Check the format.")
            st.stop()

        progress.progress(60, text=f"Parsed {len(entries)} entries.")
        status.info(f"📖 Found {len(entries)} ECN entries in {txt_file.name}.")

        # Register source
        version_str = f"{txt_version.strip()}_{txt_language.lower()}"
        existing = run_query(
            "SELECT id FROM data_sources WHERE source_type='annex_i_text' AND version=:v",
            {"v": version_str},
        )
        if existing:
            source_id = existing[0]["id"]
            execute(
                "DELETE FROM manual_entries WHERE source_id=:sid",
                {"sid": source_id},
            )
            execute(
                "UPDATE data_sources SET row_count=:rc WHERE id=:sid",
                {"rc": len(entries), "sid": source_id},
            )
            status.info(f"♻️ Replacing rows of existing source #{source_id}.")
        else:
            inserted = run_query(
                """
                INSERT INTO data_sources (source_type, source_name, version, row_count, notes)
                VALUES ('annex_i_text', :n, :v, :rc, :notes)
                RETURNING id
                """,
                {
                    "n": f"EU Annex I text ({txt_language} — {txt_version.strip()})",
                    "v": version_str,
                    "rc": len(entries),
                    "notes": f"Parsed from {txt_file.name}",
                },
            )
            source_id = inserted[0]["id"]
            status.info(f"✨ Registered new source #{source_id}.")

        # Bulk insert into manual_entries
        progress.progress(80, text=f"Inserting {len(entries)} rows...")
        import time
        from psycopg2.extras import execute_values
        from db.connection import get_engine
        import json as _json

        batch = []
        for e in entries:
            payload = {
                "code": e["code"],
                "label": e["label"],
                "category": e["category"],
                "subgroup": e["subgroup"],
                "depth": e["depth"],
                "language": txt_language,
                "regulation_version": txt_version.strip(),
            }
            batch.append({
                "source_id": source_id,
                "entry_key": e["code"],
                "entry_label": e["label"],
                "payload": _json.dumps(payload, default=str),
            })

        t0 = time.time()
        engine = get_engine()
        with engine.begin() as conn:
            raw_conn = conn.connection
            cur = raw_conn.cursor()
            execute_values(
                cur,
                "INSERT INTO manual_entries (source_id, entry_key, entry_label, payload) "
                "VALUES %s",
                batch,
                template="(%(source_id)s, %(entry_key)s, %(entry_label)s, %(payload)s::jsonb)",
                page_size=500,
            )
            cur.close()
        elapsed = time.time() - t0

        progress.progress(100, text="Done.")
        status.success(
            f"✓ Loaded {len(batch)} {txt_language} regulation entries in {elapsed:.1f}s "
            f"(source #{source_id}, version {version_str})."
        )
        time.sleep(1.5)
        st.rerun()

    except Exception as exc:
        progress.empty()
        status.empty()
        st.error(f"Load failed: {exc}")
        st.exception(exc)


st.divider()

# ----------------------------------------------------------------------
# Section 3: Upload generic CSV/JSON to manual_entries
# ----------------------------------------------------------------------
st.subheader("📥 Add a custom data source (CSV / JSON)")
st.caption(
    "Use this to upload extras like a country-risk list, a CN-to-ECN snippet "
    "you collected manually, or internal compliance notes. "
    "Each row becomes a `manual_entries` record, queryable from the "
    "**Search product** page."
)

custom_file = st.file_uploader("File", type=["csv", "json"], key="custom_upload")

with st.form("custom_source_form"):
    c1, c2, c3 = st.columns(3)
    with c1:
        source_type = st.text_input(
            "Source type",
            placeholder="country_risk, cn_eccn_map, ...",
        )
    with c2:
        source_name = st.text_input("Source name", placeholder="My BE country risk list")
    with c3:
        version = st.text_input("Version", placeholder="2025-Q4 or v1")

    key_column = st.text_input(
        "Key column (looked-up term)",
        placeholder="e.g. 'country_code' or 'cn_code'",
        help="The column that becomes `entry_key`. Used for ILIKE searches.",
    )
    label_column = st.text_input(
        "Label column (optional)",
        placeholder="e.g. 'description'",
    )
    notes = st.text_area("Notes (optional)")
    submitted = st.form_submit_button("Load custom source", type="primary")

if submitted:
    if not custom_file:
        st.warning("Upload a CSV or JSON file first.")
        st.stop()
    if not (source_type and source_name and version and key_column):
        st.warning("Source type, name, version, and key column are all required.")
        st.stop()

    try:
        raw = custom_file.getvalue()
        if custom_file.name.lower().endswith(".csv"):
            df = pd.read_csv(io.BytesIO(raw))
        else:
            df = pd.read_json(io.BytesIO(raw))

        if key_column not in df.columns:
            st.error(f"Column `{key_column}` not found. Available: {list(df.columns)}")
            st.stop()

        st.write(f"Loaded {len(df)} rows. Sample:")
        st.dataframe(df.head(5), use_container_width=True)

        inserted = run_query(
            """
            INSERT INTO data_sources (source_type, source_name, version, row_count, notes)
            VALUES (:t, :n, :v, :rc, :notes)
            RETURNING id
            """,
            {"t": source_type, "n": source_name, "v": version,
             "rc": len(df), "notes": notes},
        )
        source_id = inserted[0]["id"]

        batch = []
        for _, row in df.iterrows():
            key_val = str(row[key_column])
            label_val = (
                str(row[label_column])
                if label_column and label_column in df.columns
                else key_val
            )
            batch.append({
                "source_id": source_id,
                "entry_key": key_val,
                "entry_label": label_val,
                "payload": json.dumps(row.to_dict(), default=str),
            })

        # JSONB cast via SQL template
        from psycopg2.extras import execute_values
        from db.connection import get_engine
        engine = get_engine()
        with engine.begin() as conn:
            raw_conn = conn.connection
            cur = raw_conn.cursor()
            execute_values(
                cur,
                "INSERT INTO manual_entries (source_id, entry_key, entry_label, payload) "
                "VALUES %s",
                batch,
                template="(%(source_id)s, %(entry_key)s, %(entry_label)s, %(payload)s::jsonb)",
                page_size=500,
            )
            cur.close()

        st.success(f"✓ Loaded {len(batch)} rows from {custom_file.name} as source #{source_id}.")
        time.sleep(1.5)
        st.rerun()
    except Exception as exc:
        st.error(f"Load failed: {exc}")
        st.exception(exc)


st.divider()

# ----------------------------------------------------------------------
# Section 4: Deactivate / delete a source
# ----------------------------------------------------------------------
st.subheader("🗑️ Manage existing sources")
sources = run_query(
    "SELECT id, source_type, source_name, version, is_active "
    "FROM data_sources ORDER BY loaded_at DESC"
)
if sources:
    options = {
        f"#{s['id']} — {s['source_name']} ({s['version']}) — active={s['is_active']}": s["id"]
        for s in sources
    }
    chosen = st.selectbox("Pick a source", list(options.keys()))
    c1, c2 = st.columns(2)
    with c1:
        if st.button("Toggle active", use_container_width=True):
            execute(
                "UPDATE data_sources SET is_active = NOT is_active WHERE id = :sid",
                {"sid": options[chosen]},
            )
            st.rerun()
    with c2:
        confirm = st.checkbox("I confirm deletion", value=False)
        if st.button(
            "Delete source (and its rows)",
            use_container_width=True,
            disabled=not confirm,
        ):
            execute(
                "DELETE FROM data_sources WHERE id = :sid",
                {"sid": options[chosen]},
            )
            st.rerun()
