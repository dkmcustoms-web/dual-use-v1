"""Manage data sources: view what's loaded, add new CSV/JSON sources, load Annex I Excel."""
from __future__ import annotations

# --- Ensure project root is on sys.path (Streamlit/Railway compat) -----
import sys
from pathlib import Path
_ROOT = Path(__file__).resolve().parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))
# ----------------------------------------------------------------------

import io
import json
import os
import time

import pandas as pd
import streamlit as st

from db.connection import execute, run_query, bulk_insert_values

st.set_page_config(page_title="Data sources", page_icon="⚙️", layout="wide")
st.title("⚙️ Data sources")
st.caption(
    "Every dataset used by the sandbox is tracked here for reproducibility. "
    "Each source has a name, version, and load timestamp."
)


# ----------------------------------------------------------------------
# Current sources table
# ----------------------------------------------------------------------
st.subheader("Currently registered sources")
try:
    rows = run_query(
        """
        SELECT  id, source_type, source_name, version, loaded_at,
                row_count, is_active, notes
        FROM    data_sources
        ORDER BY loaded_at DESC
        """
    )
    if not rows:
        st.info("No sources loaded yet. Use the upload sections below.")
    else:
        df = pd.DataFrame(rows)
        st.dataframe(df, hide_index=True, use_container_width=True)
except Exception as exc:
    st.error(f"Could not read data_sources: {exc}")
    st.info("Verify DATABASE_URL is set and `python -m db.init_db` has been run.")


st.divider()

# ----------------------------------------------------------------------
# Section 1: Upload the Annex I Excel
# ----------------------------------------------------------------------
st.subheader("📚 Load Annex I from DG TRADE Excel")
st.caption(
    "Excel with columns `ID`, `PARENT_ID`, `CODE`, `LABEL` on the "
    "**Export Worksheet** sheet — as published by DG TRADE."
)

col_file, col_version = st.columns([2, 1])
with col_file:
    excel_file = st.file_uploader(
        "Annex I Excel (.xlsx)",
        type=["xlsx"],
        key="annex_upload",
    )
with col_version:
    version_label = st.text_input(
        "Version label",
        value="2024-09",
        placeholder="2024-09",
        key="annex_version",
        help="The DG TRADE Excel currently available is the September 2024 version. "
             "Change this when you load a newer version.",
    )

# Button is ALWAYS rendered, just disabled until prerequisites are met.
ready = (excel_file is not None) and bool(version_label.strip())
load_clicked = st.button(
    "🚀 Load Annex I into database" if ready else "Upload a file and enter a version first",
    type="primary",
    disabled=not ready,
    use_container_width=True,
    key="load_annex_btn",
)

if load_clicked:
    progress = st.progress(0, text="Reading Excel...")
    status = st.empty()
    try:
        import openpyxl

        # Use getvalue() — it does NOT consume the stream like read() does.
        file_bytes = excel_file.getvalue()
        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes),
            data_only=True,
            read_only=True,
        )
        if "Export Worksheet" not in wb.sheetnames:
            st.error(
                f"Sheet 'Export Worksheet' not found. "
                f"Available sheets: {wb.sheetnames}"
            )
            st.stop()
        ws = wb["Export Worksheet"]
        raw_rows = list(ws.iter_rows(min_row=2, values_only=True))
        progress.progress(20, text=f"Parsed {len(raw_rows)} rows from Excel.")
        status.info(f"📖 Read {len(raw_rows):,} rows from {excel_file.name}.")

        # ---- Register source row ------------------------------------
        progress.progress(30, text="Registering source...")
        version_str = version_label.strip()
        existing = run_query(
            "SELECT id FROM data_sources WHERE source_type='annex_i' AND version=:v",
            {"v": version_str},
        )
        if existing:
            source_id = existing[0]["id"]
            execute(
                "DELETE FROM annex_i_items WHERE source_id=:sid",
                {"sid": source_id},
            )
            status.info(f"♻️ Replacing existing source #{source_id} ({version_str}).")
        else:
            inserted = run_query(
                """
                INSERT INTO data_sources (source_type, source_name, version, row_count, notes)
                VALUES ('annex_i', :n, :v, :rc, :notes)
                RETURNING id
                """,
                {
                    "n": f"EU Annex I Dual-Use ({version_str})",
                    "v": version_str,
                    "rc": len(raw_rows),
                    "notes": f"Uploaded via UI: {excel_file.name}",
                },
            )
            source_id = inserted[0]["id"]
            execute(
                """
                UPDATE data_sources SET is_active = FALSE
                WHERE  source_type = 'annex_i' AND id <> :sid
                """,
                {"sid": source_id},
            )
            status.info(f"✨ Registered new source #{source_id} ({version_str}).")

        # ---- Build batch --------------------------------------------
        progress.progress(50, text="Preparing rows for insert...")
        batch = []
        for r in raw_rows:
            item_id, parent_id, code, label = r
            if item_id is None or code is None:
                continue
            code_str = str(code).strip()
            category = code_str[0] if code_str and code_str[0].isdigit() else None
            subgroup = code_str[1] if (category and len(code_str) >= 2 and code_str[1] in "ABCDE") else None
            depth = code_str.count(".")
            if parent_id in ("", None):
                pid_val = None
            else:
                try:
                    pid_val = int(parent_id)
                except (TypeError, ValueError):
                    pid_val = None
            batch.append({
                "id": int(item_id),
                "parent_id": pid_val,
                "code": code_str,
                "label": str(label) if label is not None else "",
                "category": category,
                "subgroup": subgroup,
                "depth": depth,
                "source_id": source_id,
            })

        # ---- Fast bulk insert via psycopg2.execute_values -----------
        progress.progress(70, text=f"Inserting {len(batch):,} rows into Postgres...")
        t0 = time.time()
        bulk_insert_values(
            table="annex_i_items",
            columns=["id", "parent_id", "code", "label",
                     "category", "subgroup", "depth", "source_id"],
            rows=batch,
            page_size=500,
        )
        elapsed = time.time() - t0

        execute(
            "UPDATE data_sources SET row_count = :rc WHERE id = :sid",
            {"rc": len(batch), "sid": source_id},
        )

        progress.progress(100, text="Done.")
        status.success(
            f"✓ Loaded {len(batch):,} Annex I rows in {elapsed:.1f}s "
            f"(source #{source_id}, version {version_str})."
        )
        time.sleep(1.5)
        st.rerun()

    except Exception as exc:
        progress.empty()
        status.empty()
        st.error(f"Load failed: {exc}")
        st.exception(exc)


st.divider()

# ----------------------------------------------------------------------
# Section 2: Load BUNDLED Dutch Annex I (no upload needed)
# ----------------------------------------------------------------------
st.subheader("🇳🇱 Import bundled Dutch Annex I (2025-11)")
st.caption(
    "Pre-parsed JSON of the consolidated Dutch text "
    "(EUR-Lex 02021R0821-NL-15.11.2025, ~400 entries with full narrative content). "
    "Lives in `data/annex_i_nl_2025-11.json` — no upload required."
)

bundled_nl_path = _ROOT / "data" / "annex_i_nl_2025-11.json"
bundled_en_path = _ROOT / "data" / "annex_i_en_2025-11.json"


def _import_bundled(path, language_label, language_code):
    """Import a bundled language JSON into manual_entries. Reusable for NL + EN."""
    progress = st.progress(0, text=f"Reading bundled {language_label} JSON...")
    status = st.empty()
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        meta = data.get("metadata", {})
        entries = data.get("entries", [])
        progress.progress(30, text=f"Read {len(entries)} {language_label} entries.")
        status.info(
            f"📖 Bundled JSON: {meta.get('source')} — "
            f"v{meta.get('version')} — {len(entries)} entries."
        )

        version_str = f"{meta.get('version', '2025-11')}_{language_code.lower()}_bundled"
        existing = run_query(
            "SELECT id FROM data_sources WHERE source_type='annex_i_text' AND version=:v",
            {"v": version_str},
        )
        if existing:
            source_id = existing[0]["id"]
            execute("DELETE FROM manual_entries WHERE source_id=:sid", {"sid": source_id})
            execute(
                "UPDATE data_sources SET row_count=:rc WHERE id=:sid",
                {"rc": len(entries), "sid": source_id},
            )
            status.info(f"♻️ Replacing rows of existing source #{source_id}.")
        else:
            inserted = run_query(
                """
                INSERT INTO data_sources (source_type, source_name, version, row_count, notes)
                VALUES ('annex_i_text', :n, :v, :rc, :notes)
                RETURNING id
                """,
                {
                    "n": f"EU Annex I ({language_code}) — {meta.get('version', '2025-11')} bundled",
                    "v": version_str,
                    "rc": len(entries),
                    "notes": f"Loaded from bundled JSON (data/annex_i_{language_code.lower()}_2025-11.json).",
                },
            )
            source_id = inserted[0]["id"]

        progress.progress(60, text=f"Inserting {len(entries)} rows...")
        from psycopg2.extras import execute_values
        from db.connection import get_engine
        batch = []
        for e in entries:
            payload = {
                "code": e.get("code"),
                "label": e.get("label"),
                "category": e.get("category"),
                "subgroup": e.get("subgroup"),
                "depth": e.get("depth"),
                "full_content": e.get("full_content", ""),
                "language": language_code,
                "regulation_version": meta.get("version"),
                "source": "bundled",
            }
            batch.append({
                "source_id": source_id,
                "entry_key": e.get("code"),
                "entry_label": e.get("label"),
                "payload": json.dumps(payload, default=str),
            })
        t0 = time.time()
        engine = get_engine()
        with engine.begin() as conn:
            raw_conn = conn.connection
            cur = raw_conn.cursor()
            execute_values(
                cur,
                "INSERT INTO manual_entries (source_id, entry_key, entry_label, payload) VALUES %s",
                batch,
                template="(%(source_id)s, %(entry_key)s, %(entry_label)s, %(payload)s::jsonb)",
                page_size=500,
            )
            cur.close()
        elapsed = time.time() - t0

        progress.progress(100, text="Done.")
        status.success(
            f"✓ Imported {len(batch)} {language_label} entries in {elapsed:.1f}s (source #{source_id})."
        )
        time.sleep(1.5)
        st.rerun()
    except Exception as exc:
        progress.empty()
        status.empty()
        st.error(f"Import failed: {exc}")
        st.exception(exc)


if not bundled_nl_path.exists():
    st.warning(f"Bundled NL file not found at {bundled_nl_path}.")
else:
    if st.button("📦 Import bundled NL Annex I (2025-11)", type="primary", use_container_width=True, key="import_bundled_nl_btn"):
        _import_bundled(bundled_nl_path, "Dutch", "NL")

st.divider()

# ----------------------------------------------------------------------
# Section 3: Load BUNDLED English Annex I
# ----------------------------------------------------------------------
st.subheader("🇬🇧 Import bundled English Annex I (2025-11)")
st.caption(
    "Pre-parsed JSON of the consolidated English text "
    "(EUR-Lex 02021R0821-EN-15.11.2025, ~550 entries with full narrative content). "
    "Combined with NL gives bilingual narrative coverage for cross-language search."
)
if not bundled_en_path.exists():
    st.warning(f"Bundled EN file not found at {bundled_en_path}.")
else:
    if st.button("📦 Import bundled EN Annex I (2025-11)", type="primary", use_container_width=True, key="import_bundled_en_btn"):
        _import_bundled(bundled_en_path, "English", "EN")

st.divider()

# ----------------------------------------------------------------------
# Section 4: Build embeddings for semantic search
# ----------------------------------------------------------------------
st.subheader("🧠 Build semantic search index")
st.caption(
    "Embeds all loaded entries with OpenAI `text-embedding-3-small` (1536 dim, "
    "multilingual). One-time cost ≈ $0.01–0.02. Per-query cost ≈ $0.0000001 (effectively free)."
)

from services import annex_i as _ai
from services import embeddings as _emb

annex_stats = _ai.count_embedded_annex_items()
manual_stats = _ai.count_embedded_manual_entries()

c1, c2, c3 = st.columns(3)
with c1:
    st.metric(
        "Annex I (Excel)",
        f"{annex_stats['embedded']}/{annex_stats['total']}",
        delta=f"{annex_stats['remaining']} to embed" if annex_stats['remaining'] else "✓ done",
    )
with c2:
    st.metric(
        "Manual entries (NL/EN TXT)",
        f"{manual_stats['embedded']}/{manual_stats['total']}",
        delta=f"{manual_stats['remaining']} to embed" if manual_stats['remaining'] else "✓ done",
    )
with c3:
    total_remaining = annex_stats['remaining'] + manual_stats['remaining']
    if total_remaining:
        est = _emb.estimate_cost_for_corpus(total_remaining)
        st.metric("Est. cost", f"${est['estimated_cost_usd']:.4f}")
    else:
        st.metric("Est. cost", "$0")

if not os.environ.get("OPENAI_API_KEY"):
    st.error("🔑 `OPENAI_API_KEY` is not set. Add it to .env locally or to Railway → Variables.")
else:
    cols = st.columns(2)
    with cols[0]:
        build_annex = st.button(
            f"🧠 Embed Annex I ({annex_stats['remaining']} remaining)",
            disabled=annex_stats['remaining'] == 0,
            use_container_width=True,
            key="embed_annex_btn",
        )
    with cols[1]:
        build_manual = st.button(
            f"🧠 Embed manual entries ({manual_stats['remaining']} remaining)",
            disabled=manual_stats['remaining'] == 0,
            use_container_width=True,
            key="embed_manual_btn",
        )

    if build_annex:
        rows = run_query(
            """
            SELECT a.id, a.code, a.label
            FROM   annex_i_items a
            JOIN   data_sources ds ON ds.id = a.source_id
            WHERE  ds.is_active = TRUE
              AND  a.code ~ '^[0-9][A-E][0-9]{3}'
              AND  a.embedding IS NULL
            ORDER BY a.id
            """
        )
        if rows:
            progress = st.progress(0, text=f"Embedding {len(rows)} Annex I rows...")
            status = st.empty()
            try:
                texts = [
                    f"{r['code']} {r['label']}".strip()
                    for r in rows
                ]
                BATCH = 200
                total_tokens = 0
                total_cost = 0.0
                t0 = time.time()
                for i in range(0, len(rows), BATCH):
                    chunk_rows = rows[i:i + BATCH]
                    chunk_texts = texts[i:i + BATCH]
                    progress.progress(
                        min(0.99, (i + 1) / len(rows)),
                        text=f"Embedding batch {i // BATCH + 1} / {(len(rows) - 1) // BATCH + 1}...",
                    )
                    result = _emb.embed_texts(chunk_texts)
                    total_tokens += result["total_tokens"]
                    total_cost += result["cost_usd"]

                    # Update DB
                    update_rows = []
                    for r, vec in zip(chunk_rows, result["embeddings"]):
                        update_rows.append({
                            "id": r["id"],
                            "vec": _emb.to_pg_vector_literal(vec),
                        })
                    for ur in update_rows:
                        execute(
                            """
                            UPDATE annex_i_items
                            SET embedding = CAST(:vec AS vector),
                                embedding_model = :m,
                                embedded_at = NOW()
                            WHERE id = :id
                            """,
                            {"vec": ur["vec"], "m": _emb.EMBEDDING_MODEL, "id": ur["id"]},
                        )

                progress.progress(1.0, text="Done.")
                elapsed = time.time() - t0
                status.success(
                    f"✓ Embedded {len(rows)} rows in {elapsed:.1f}s. "
                    f"Tokens used: {total_tokens:,}. Cost: ${total_cost:.4f}."
                )
                time.sleep(2)
                st.rerun()
            except Exception as exc:
                progress.empty()
                st.error(f"Embedding failed: {exc}")
                st.exception(exc)

    if build_manual:
        rows = run_query(
            """
            SELECT  me.entry_key AS code,
                    me.entry_label AS label,
                    me.payload AS payload,
                    me.source_id || '_' || me.entry_key AS uid,
                    me.source_id,
                    me.entry_key
            FROM    manual_entries me
            JOIN    data_sources ds ON ds.id = me.source_id
            WHERE   ds.is_active = TRUE
              AND   me.embedding IS NULL
            ORDER BY me.source_id, me.entry_key
            """
        )
        if rows:
            progress = st.progress(0, text=f"Embedding {len(rows)} manual entries...")
            status = st.empty()
            try:
                # Build text from code + label + full_content (capped for cost control)
                texts = []
                for r in rows:
                    p = r.get("payload") or {}
                    if isinstance(p, str):
                        try:
                            p = json.loads(p)
                        except Exception:
                            p = {}
                    full = (p.get("full_content") or "")[:2000] if isinstance(p, dict) else ""
                    t = f"{r['code']} {r['label']} {full}".strip()
                    texts.append(t)

                BATCH = 100
                total_tokens = 0
                total_cost = 0.0
                t0 = time.time()
                for i in range(0, len(rows), BATCH):
                    chunk_rows = rows[i:i + BATCH]
                    chunk_texts = texts[i:i + BATCH]
                    progress.progress(
                        min(0.99, (i + 1) / len(rows)),
                        text=f"Embedding batch {i // BATCH + 1} / {(len(rows) - 1) // BATCH + 1}...",
                    )
                    result = _emb.embed_texts(chunk_texts)
                    total_tokens += result["total_tokens"]
                    total_cost += result["cost_usd"]

                    for r, vec in zip(chunk_rows, result["embeddings"]):
                        execute(
                            """
                            UPDATE manual_entries
                            SET embedding = CAST(:vec AS vector),
                                embedding_model = :m,
                                embedded_at = NOW()
                            WHERE source_id = :sid AND entry_key = :ek
                            """,
                            {
                                "vec": _emb.to_pg_vector_literal(vec),
                                "m": _emb.EMBEDDING_MODEL,
                                "sid": r["source_id"],
                                "ek": r["entry_key"],
                            },
                        )

                progress.progress(1.0, text="Done.")
                elapsed = time.time() - t0
                status.success(
                    f"✓ Embedded {len(rows)} rows in {elapsed:.1f}s. "
                    f"Tokens: {total_tokens:,}. Cost: ${total_cost:.4f}."
                )
                time.sleep(2)
                st.rerun()
            except Exception as exc:
                progress.empty()
                st.error(f"Embedding failed: {exc}")
                st.exception(exc)


st.divider()

# ----------------------------------------------------------------------
# Section 5: Upload Dutch (or other language) regulation text MANUALLY
# ----------------------------------------------------------------------
st.subheader("📄 Load regulation text (Dual_use.txt — Dutch / EN / FR / DE)")
st.caption(
    "Plaintext export of the consolidated EU regulation from EUR-Lex. "
    "Extracts ECN code → first-line label pairs and stores them as "
    "`manual_entries` with source_type `annex_i_text`. "
    "Searchable from the **Search product** page alongside the structured Excel."
)

col_txt, col_lang, col_ver = st.columns([2, 1, 1])
with col_txt:
    txt_file = st.file_uploader(
        "Regulation text file (.txt)",
        type=["txt"],
        key="regulation_txt_upload",
    )
with col_lang:
    txt_language = st.selectbox(
        "Language",
        options=["NL", "EN", "FR", "DE", "other"],
        index=0,
        key="regulation_txt_lang",
    )
with col_ver:
    txt_version = st.text_input(
        "Version",
        value="2025-11",
        placeholder="2025-11",
        key="regulation_txt_version",
    )

txt_ready = (txt_file is not None) and bool(txt_version.strip())
txt_clicked = st.button(
    "🚀 Parse and load regulation text" if txt_ready
    else "Upload a .txt and enter a version first",
    type="primary",
    disabled=not txt_ready,
    use_container_width=True,
    key="load_txt_btn",
)

if txt_clicked:
    progress = st.progress(0, text="Reading file...")
    status = st.empty()
    try:
        from services.regulation_parser import parse_eu_regulation_txt

        # Decode bytes → string (try UTF-8 first, fall back to latin-1)
        raw = txt_file.getvalue()
        try:
            text_content = raw.decode("utf-8")
        except UnicodeDecodeError:
            text_content = raw.decode("latin-1")
        progress.progress(20, text=f"Read {len(text_content):,} chars.")

        # Parse
        progress.progress(40, text="Extracting ECN entries...")
        entries = parse_eu_regulation_txt(text_content)
        if not entries:
            st.warning("No ECN-style entries found in the file. Check the format.")
            st.stop()

        progress.progress(60, text=f"Parsed {len(entries)} entries.")
        status.info(f"📖 Found {len(entries)} ECN entries in {txt_file.name}.")

        # Register source
        version_str = f"{txt_version.strip()}_{txt_language.lower()}"
        existing = run_query(
            "SELECT id FROM data_sources WHERE source_type='annex_i_text' AND version=:v",
            {"v": version_str},
        )
        if existing:
            source_id = existing[0]["id"]
            execute(
                "DELETE FROM manual_entries WHERE source_id=:sid",
                {"sid": source_id},
            )
            execute(
                "UPDATE data_sources SET row_count=:rc WHERE id=:sid",
                {"rc": len(entries), "sid": source_id},
            )
            status.info(f"♻️ Replacing rows of existing source #{source_id}.")
        else:
            inserted = run_query(
                """
                INSERT INTO data_sources (source_type, source_name, version, row_count, notes)
                VALUES ('annex_i_text', :n, :v, :rc, :notes)
                RETURNING id
                """,
                {
                    "n": f"EU Annex I text ({txt_language} — {txt_version.strip()})",
                    "v": version_str,
                    "rc": len(entries),
                    "notes": f"Parsed from {txt_file.name}",
                },
            )
            source_id = inserted[0]["id"]
            status.info(f"✨ Registered new source #{source_id}.")

        # Bulk insert into manual_entries
        progress.progress(80, text=f"Inserting {len(entries)} rows...")
        import time
        from psycopg2.extras import execute_values
        from db.connection import get_engine
        import json as _json

        batch = []
        for e in entries:
            payload = {
                "code": e["code"],
                "label": e["label"],
                "category": e["category"],
                "subgroup": e["subgroup"],
                "depth": e["depth"],
                "full_content": e.get("full_content", ""),
                "language": txt_language,
                "regulation_version": txt_version.strip(),
            }
            batch.append({
                "source_id": source_id,
                "entry_key": e["code"],
                "entry_label": e["label"],
                "payload": _json.dumps(payload, default=str),
            })

        t0 = time.time()
        engine = get_engine()
        with engine.begin() as conn:
            raw_conn = conn.connection
            cur = raw_conn.cursor()
            execute_values(
                cur,
                "INSERT INTO manual_entries (source_id, entry_key, entry_label, payload) "
                "VALUES %s",
                batch,
                template="(%(source_id)s, %(entry_key)s, %(entry_label)s, %(payload)s::jsonb)",
                page_size=500,
            )
            cur.close()
        elapsed = time.time() - t0

        progress.progress(100, text="Done.")
        status.success(
            f"✓ Loaded {len(batch)} {txt_language} regulation entries in {elapsed:.1f}s "
            f"(source #{source_id}, version {version_str})."
        )
        time.sleep(1.5)
        st.rerun()

    except Exception as exc:
        progress.empty()
        status.empty()
        st.error(f"Load failed: {exc}")
        st.exception(exc)


st.divider()

# ----------------------------------------------------------------------
# Section 3: Upload generic CSV/JSON to manual_entries
# ----------------------------------------------------------------------
st.subheader("📥 Add a custom data source (CSV / JSON)")
st.caption(
    "Use this to upload extras like a country-risk list, a CN-to-ECN snippet "
    "you collected manually, or internal compliance notes. "
    "Each row becomes a `manual_entries` record, queryable from the "
    "**Search product** page."
)

custom_file = st.file_uploader("File", type=["csv", "json"], key="custom_upload")

with st.form("custom_source_form"):
    c1, c2, c3 = st.columns(3)
    with c1:
        source_type = st.text_input(
            "Source type",
            placeholder="country_risk, cn_eccn_map, ...",
        )
    with c2:
        source_name = st.text_input("Source name", placeholder="My BE country risk list")
    with c3:
        version = st.text_input("Version", placeholder="2025-Q4 or v1")

    key_column = st.text_input(
        "Key column (looked-up term)",
        placeholder="e.g. 'country_code' or 'cn_code'",
        help="The column that becomes `entry_key`. Used for ILIKE searches.",
    )
    label_column = st.text_input(
        "Label column (optional)",
        placeholder="e.g. 'description'",
    )
    notes = st.text_area("Notes (optional)")
    submitted = st.form_submit_button("Load custom source", type="primary")

if submitted:
    if not custom_file:
        st.warning("Upload a CSV or JSON file first.")
        st.stop()
    if not (source_type and source_name and version and key_column):
        st.warning("Source type, name, version, and key column are all required.")
        st.stop()

    try:
        raw = custom_file.getvalue()
        if custom_file.name.lower().endswith(".csv"):
            df = pd.read_csv(io.BytesIO(raw))
        else:
            df = pd.read_json(io.BytesIO(raw))

        if key_column not in df.columns:
            st.error(f"Column `{key_column}` not found. Available: {list(df.columns)}")
            st.stop()

        st.write(f"Loaded {len(df)} rows. Sample:")
        st.dataframe(df.head(5), use_container_width=True)

        inserted = run_query(
            """
            INSERT INTO data_sources (source_type, source_name, version, row_count, notes)
            VALUES (:t, :n, :v, :rc, :notes)
            RETURNING id
            """,
            {"t": source_type, "n": source_name, "v": version,
             "rc": len(df), "notes": notes},
        )
        source_id = inserted[0]["id"]

        batch = []
        for _, row in df.iterrows():
            key_val = str(row[key_column])
            label_val = (
                str(row[label_column])
                if label_column and label_column in df.columns
                else key_val
            )
            batch.append({
                "source_id": source_id,
                "entry_key": key_val,
                "entry_label": label_val,
                "payload": json.dumps(row.to_dict(), default=str),
            })

        # JSONB cast via SQL template
        from psycopg2.extras import execute_values
        from db.connection import get_engine
        engine = get_engine()
        with engine.begin() as conn:
            raw_conn = conn.connection
            cur = raw_conn.cursor()
            execute_values(
                cur,
                "INSERT INTO manual_entries (source_id, entry_key, entry_label, payload) "
                "VALUES %s",
                batch,
                template="(%(source_id)s, %(entry_key)s, %(entry_label)s, %(payload)s::jsonb)",
                page_size=500,
            )
            cur.close()

        st.success(f"✓ Loaded {len(batch)} rows from {custom_file.name} as source #{source_id}.")
        time.sleep(1.5)
        st.rerun()
    except Exception as exc:
        st.error(f"Load failed: {exc}")
        st.exception(exc)


st.divider()

# ----------------------------------------------------------------------
# Section 4: Compliance prompts (for the AI review page)
# ----------------------------------------------------------------------
st.subheader("🤖 Compliance prompts (for AI review)")
st.caption(
    "System prompts used on the **AI compliance review** page. Multiple "
    "versions can coexist; only one is active at any time. "
    "Edit, version, and tune them here without touching code."
)

from services import llm_review as _llmr  # noqa: E402

prompts = _llmr.list_prompts()
if not prompts:
    st.info("No compliance prompts yet. Seed the default EU compliance prompt below.")
    if st.button("⚡ Seed default EU compliance prompt", type="primary"):
        try:
            new_id = _llmr.seed_default_prompt()
            st.success(f"✓ Default prompt seeded as #{new_id}.")
            time.sleep(1)
            st.rerun()
        except Exception as exc:
            st.error(f"Seed failed: {exc}")
else:
    df_prompts = pd.DataFrame(prompts)
    st.dataframe(df_prompts, hide_index=True, use_container_width=True)

    cA, cB = st.columns(2)
    with cA:
        chosen_id = st.selectbox(
            "Pick a prompt to view / edit",
            options=[p["id"] for p in prompts],
            format_func=lambda i: next(
                f"#{p['id']} — {p['name']} (v{p['version']}) {'★' if p['is_active'] else ''}"
                for p in prompts if p["id"] == i
            ),
        )
    with cB:
        c1, c2 = st.columns(2)
        with c1:
            if st.button("Activate this one", use_container_width=True):
                _llmr.activate_prompt(chosen_id)
                st.rerun()
        with c2:
            if st.button(
                "🗑️ Delete prompt",
                use_container_width=True,
                disabled=len(prompts) <= 1,
                help="Cannot delete the last remaining prompt.",
            ):
                _llmr.delete_prompt(chosen_id)
                st.rerun()

    full = _llmr.get_prompt(chosen_id)
    if full:
        with st.form("edit_prompt_form"):
            st.markdown(f"### Edit prompt #{full['id']}")
            ec1, ec2, ec3 = st.columns([2, 1, 1])
            with ec1:
                e_name = st.text_input("Name", value=full["name"])
            with ec2:
                e_version = st.text_input("Version", value=full["version"])
            with ec3:
                e_model = st.text_input(
                    "Model",
                    value=full["model"],
                    help="Anthropic model id, e.g. claude-sonnet-4-6, "
                         "claude-opus-4-7, claude-haiku-4-5-20251001",
                )
            ec4, ec5 = st.columns(2)
            with ec4:
                e_temp = st.number_input(
                    "Temperature",
                    min_value=0.0,
                    max_value=1.0,
                    step=0.1,
                    value=float(full["temperature"]),
                )
            with ec5:
                e_max_tokens = st.number_input(
                    "Max tokens",
                    min_value=500,
                    max_value=16000,
                    step=500,
                    value=int(full["max_tokens"]),
                )
            e_notes = st.text_input("Notes", value=full.get("notes") or "")
            e_system = st.text_area(
                "System prompt",
                value=full["system_prompt"],
                height=400,
            )
            save_as_new = st.checkbox(
                "Save as new version (creates a new prompt row; original kept)",
                value=False,
            )

            saved = st.form_submit_button("💾 Save", type="primary")

        if saved:
            try:
                target_version = e_version
                if save_as_new and target_version == full["version"]:
                    target_version = f"{full['version']}-edit-{int(time.time())}"
                new_id = _llmr.save_prompt(
                    name=e_name,
                    version=target_version,
                    system_prompt=e_system,
                    model=e_model,
                    temperature=float(e_temp),
                    max_tokens=int(e_max_tokens),
                    notes=e_notes,
                    activate=True,
                )
                st.success(f"✓ Saved as prompt #{new_id} (version {target_version}). Activated.")
                time.sleep(1)
                st.rerun()
            except Exception as exc:
                st.error(f"Save failed: {exc}")


st.divider()

# ----------------------------------------------------------------------
# Section 5: Deactivate / delete a data source
# ----------------------------------------------------------------------
st.subheader("🗑️ Manage existing sources")
sources = run_query(
    "SELECT id, source_type, source_name, version, is_active "
    "FROM data_sources ORDER BY loaded_at DESC"
)
if sources:
    options = {
        f"#{s['id']} — {s['source_name']} ({s['version']}) — active={s['is_active']}": s["id"]
        for s in sources
    }
    chosen = st.selectbox("Pick a source", list(options.keys()))
    c1, c2 = st.columns(2)
    with c1:
        if st.button("Toggle active", use_container_width=True):
            execute(
                "UPDATE data_sources SET is_active = NOT is_active WHERE id = :sid",
                {"sid": options[chosen]},
            )
            st.rerun()
    with c2:
        confirm = st.checkbox("I confirm deletion", value=False)
        if st.button(
            "Delete source (and its rows)",
            use_container_width=True,
            disabled=not confirm,
        ):
            execute(
                "DELETE FROM data_sources WHERE id = :sid",
                {"sid": options[chosen]},
            )
            st.rerun()
